from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import pymongo, datetime, time
import pandas as pd
import openpyxl as op
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.secret_key = "supersecretkey"  # Necesario para usar mensajes flash

# Conexión con MongoDB
try:
    client = pymongo.MongoClient("mongodb://localhost:27017/")
    db = client["rueda_vida"]
    collection = db["empleados"]
    print("✅ Conexión exitosa a MongoDB")
except Exception as e:
    print("❌ Error de conexión a MongoDB:", e)

@app.route('/')
def principal():
    return render_template('index.html')

@app.route('/nosotros')
def nosotros():
    return render_template('nosotros.html')

@app.route('/contacto')
def contacto():
    return render_template('contacto.html')

@app.route('/persona')
def persona():
    return render_template('persona.html')

@app.route('/guardar_p', methods=['POST'])
def guardar_p():
    accion = request.form.get('accion')
    id = request.form.get('id')
    nombre = request.form.get('nombre')
    apellido = request.form.get('apellido')
    correo = request.form.get('correo')
    departamento = request.form.get('departamento')
    respuestas = request.form.getlist('respuestas')
    respuestas = [int(r) for r in respuestas]

    preguntas_rueda = {
        "Área Física": respuestas[:5],
        "Área Personal": respuestas[5:10],
        "Área Familiar": respuestas[10:15],
        "Área Económica": respuestas[15:20],
        "Área Profesional": respuestas[20:25],
        "Área Social": respuestas[25:30],
        "Área de Ocio": respuestas[30:35],
        "Área Espiritual": respuestas[35:40]
    }
    
    empleado = {
        "id": id,
        "nombre": nombre,
        "apellido": apellido,
        "correo": correo,
        "departamento": departamento,
        "respuestas": preguntas_rueda
    }
    
    if accion == "Guardar":
        collection.insert_one(empleado)
    elif accion == "Editar":
        collection.update_one({"id": id}, {"$set": empleado})
    elif accion == "Borrar":
        collection.delete_one({"id": id})
    elif accion == "Exportar":
        exportar_datos()
    
    return redirect(url_for('principal'))

@app.route('/confirmar_actualizacion', methods=['POST'])
def confirmar_actualizacion():
    id = request.form.get('id')
    nombre = request.form.get('nombre')
    apellido = request.form.get('apellido')
    correo = request.form.get('correo')
    departamento = request.form.get('departamento')
    
    if not id:
        flash("⚠️ Debes proporcionar un ID válido.", "warning")
        return redirect(url_for('resultados'))
    
    resultado = collection.update_one({"id": id}, {"$set": {
        "nombre": nombre,
        "apellido": apellido,
        "correo": correo,
        "departamento": departamento
    }})
    
    if resultado.matched_count:
        flash("✅ Datos actualizados correctamente.", "success")
    else:
        flash("⚠️ No se encontró el empleado con el ID proporcionado.", "danger")
    
    return redirect(url_for('resultados'))

@app.route('/actualizar/<id>', methods=['GET'])
def actualizar(id):
    empleado = collection.find_one({"id": id})
    
    if not empleado:
        flash("⚠️ No se encontró el empleado para eliminar.", "danger")
        return redirect(url_for('resultados'))
    
    return render_template('confirmar_actualizacion.html', empleado=empleado)

@app.route('/descargar_todos', methods=['GET'], endpoint='descargar')
def descargar_todos_empleados():
    empleados = list(collection.find({}, {'_id': 0}))

    if not empleados:
        return "⚠️ No hay empleados para exportar.", 404

    # Aplanar y procesar los datos
    datos_excel = []

    for emp in empleados:
        fila = {
            "ID": emp.get("id", ""),
            "Nombre": emp.get("nombre", ""),
            "Apellido": emp.get("apellido", ""),
            "Correo": emp.get("correo", ""),
            "Departamento": emp.get("departamento", "")
        }

        total_general = 0
        areas_con_respuesta = 0

        for area, respuestas in emp.get("respuestas", {}).items():
            # Guardar respuestas individuales
            

            # Calcular promedio del área
            if respuestas:
                promedio = sum(respuestas) / (len(respuestas) * 10) * 100  # Escala 0-10
                fila[f"{area} (%)"] = round(promedio, 2)
                total_general += promedio
                areas_con_respuesta += 1
            else:
                fila[f"{area} (%)"] = 0

        # Agregar promedio total general
        fila["Promedio General (%)"] = round(total_general / areas_con_respuesta, 2) if areas_con_respuesta > 0 else 0

        datos_excel.append(fila)

    # Crear y ordenar DataFrame
    df = pd.DataFrame(datos_excel)
    df.sort_values(by="Promedio General (%)", ascending=False, inplace=True)

    # Guardar en memoria
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Empleados')

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # Estilizar encabezado
    header_font = Font(bold=True, color='FFFFFF')
    fill = PatternFill(start_color='228B22', end_color='228B22', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = alignment

    # Ajustar ancho columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Guardar nuevamente en memoria
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return send_file(
        final_output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name="empleados.xlsx"
    )




@app.route('/eliminar/<id>', methods=['GET'])
def eliminar(id):
    collection.delete_one({"id": id})
    flash("✅ Empleado eliminado correctamente.", "success")
    return redirect(url_for('resultados'))

@app.route('/resultados')
def resultados():
    empleados = collection.find()
    return render_template('resultados.html', empleados=empleados)

@app.route('/ver/<id>')
def ver_detalles(id):
    import matplotlib.pyplot as plt
    import numpy as np
    import io
    import base64
    
    empleado = collection.find_one({"id": id}, {"_id": 0})
    
    if not empleado:
        flash("⚠️ No se encontró el empleado.", "danger")
        return redirect(url_for('resultados'))
   
    respuestas_porcentajes = {
        area: round((sum(valores) / 50) * 100, 2) for area, valores in empleado["respuestas"].items()
    }
    
    fig = plt.figure(figsize=(8, 8))
    
    fig.tight_layout(pad=3.0)
   
    ax = fig.add_subplot(111, polar=True)
    
    categorias = list(respuestas_porcentajes.keys())
    valores = list(respuestas_porcentajes.values())
     
    angulos = np.linspace(0, 2*np.pi, len(categorias), endpoint=False).tolist()
    valores.append(valores[0])
    angulos.append(angulos[0])
      
    ax.plot(angulos, valores, 'o-', linewidth=2.5, color='#1f77b4')
    ax.fill(angulos, valores, alpha=0.25, color='#1f77b4')
    
    ax.set_thetagrids(np.degrees(angulos[:-1]), categorias)  
    
    ax.set_ylim(0, 100)
     
    ax.grid(True, alpha=0.3)
    
    plt.subplots_adjust(top=0.95, bottom=0.05, right=0.95, left=0.05)
    
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
    buf.seek(0)
    img_str = base64.b64encode(buf.read()).decode('utf-8')
    img_data = f'data:image/png;base64,{img_str}'
    plt.close(fig)
    
    return render_template('detalles.html', empleado=empleado, 
                          respuestas_json=respuestas_porcentajes,
                          img_data=img_data)

# Función para exportar datos a un archivo de texto
def exportar_datos():
    empleados = collection.find()
    data = "\n".join([str(emp) for emp in empleados])
    filename = f"Reports/Personas_{time.strftime('%Y%m%d_%H%M%S')}.txt"
    with open(filename, "w") as file:
        file.write(data)
    print("Datos exportados a:", filename)

if __name__ == '__main__':
    app.run(debug=True, port=5080)